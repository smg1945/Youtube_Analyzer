"""
엑셀 내보내기 전용 모듈
분석 결과를 엑셀로 내보내기, 포맷팅, 차트 생성 담당
"""

import os
import pandas as pd
from datetime import datetime
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import config

class ExcelExporter:
    """엑셀 내보내기 클래스"""
    
    def __init__(self, filename=None):
        """
        엑셀 내보내기 초기화
        
        Args:
            filename (str): 생성할 엑셀 파일명
        """
        if filename:
            self.filename = filename
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.filename = f"YouTube_Analysis_{timestamp}.xlsx"
        
        self.workbook = None
        self.worksheet = None
        
        # 스타일 정의
        self.styles = self._define_styles()
    
    def export_video_analysis(self, video_data_list, analysis_settings, include_charts=True):
        """
        영상 분석 결과를 엑셀로 내보내기
        
        Args:
            video_data_list (list): 분석된 영상 데이터 목록
            analysis_settings (dict): 분석 설정 정보
            include_charts (bool): 차트 포함 여부
            
        Returns:
            str: 생성된 파일명
        """
        try:
            print(f"📊 엑셀 생성 시작: {len(video_data_list)}개 영상")
            
            # 데이터프레임 생성
            df = self._create_video_dataframe(video_data_list)
            
            # 엑셀 파일 생성
            with pd.ExcelWriter(self.filename, engine='xlsxwriter') as writer:
                # 워크북과 워크시트 객체 가져오기
                workbook = writer.book
                
                # 1. 메인 데이터 시트
                df.to_excel(writer, sheet_name='영상 분석 결과', index=False)
                worksheet = writer.sheets['영상 분석 결과']
                self._apply_main_sheet_formatting(workbook, worksheet, df)
                
                # 2. 요약 시트
                self._create_summary_sheet(writer, video_data_list, analysis_settings)
                
                # 3. 트렌드 키워드 시트
                self._create_keywords_sheet(writer, video_data_list)
                
                # 4. 영상 유형별 분석 시트
                self._create_video_type_analysis_sheet(writer, video_data_list)
                
                # 5. 성과 분석 시트
                self._create_performance_analysis_sheet(writer, video_data_list)
            
            # 썸네일 이미지 삽입 (openpyxl 사용)
            if config.THUMBNAIL_COLUMN_WIDTH:
                self._insert_thumbnails(video_data_list)
            
            # 차트 추가
            if include_charts:
                self._add_charts(video_data_list)
            
            print(f"✅ 엑셀 파일 생성 완료: {self.filename}")
            return self.filename
            
        except Exception as e:
            print(f"❌ 엑셀 생성 오류: {e}")
            raise
    
    def _create_video_dataframe(self, video_data_list):
        """영상 데이터를 데이터프레임으로 변환"""
        data = []
        
        for video_data in video_data_list:
            snippet = video_data.get('snippet', {})
            statistics = video_data.get('statistics', {})
            content_details = video_data.get('contentDetails', {})
            analysis = video_data.get('analysis', {})
            
            # 영상 길이 정보
            duration_seconds = analysis.get('duration_seconds', 0)
            formatted_duration = analysis.get('formatted_duration', '00:00')
            video_type = analysis.get('video_type', '알수없음')
            
            # 카테고리 이름 변환
            category_id = snippet.get('categoryId', '')
            category_name = config.YOUTUBE_CATEGORIES.get(category_id, '기타')
            
            row = {
                # 기본 정보
                '순위': video_data.get('rank', 0),
                '썸네일': '',  # 이미지는 별도로 삽입
                '제목': snippet.get('title', ''),
                '채널명': snippet.get('channelTitle', ''),
                '카테고리': category_name,
                
                # 영상 정보
                '영상유형': video_type,
                '영상길이': formatted_duration,
                '영상길이_초': duration_seconds,
                '업로드일시': self._format_datetime(snippet.get('publishedAt', '')),
                
                # 성과 지표
                '조회수': int(statistics.get('viewCount', 0)),
                '좋아요': int(statistics.get('likeCount', 0)),
                '댓글수': int(statistics.get('commentCount', 0)),
                '좋아요율': analysis.get('like_rate', 0),
                '댓글율': analysis.get('comment_rate', 0),
                
                # 분석 지표
                'Outlier점수': analysis.get('outlier_score', 1.0),
                'Outlier등급': analysis.get('outlier_category', '😐 평균'),
                '참여도점수': analysis.get('engagement_score', 0),
                '일평균조회수': analysis.get('views_per_day', 0),
                '성장속도': analysis.get('growth_velocity', {}).get('velocity_rating', '알수없음'),
                
                # 채널 비교
                '채널평균조회수': analysis.get('channel_avg_views', 0),
                '채널대비성과': f"{analysis.get('outlier_score', 1.0):.1f}x",
                
                # 컨텐츠 분석
                '핵심키워드': ', '.join(analysis.get('keywords', [])),
                '제목길이': len(snippet.get('title', '')),
                
                # 감정 분석 (댓글이 있는 경우)
                '댓글감정_긍정': f"{analysis.get('sentiment', {}).get('positive', 0):.1f}%",
                '댓글감정_중립': f"{analysis.get('sentiment', {}).get('neutral', 0):.1f}%",
                '댓글감정_부정': f"{analysis.get('sentiment', {}).get('negative', 0):.1f}%",
                
                # 링크
                '영상링크': f"https://www.youtube.com/watch?v={video_data.get('id', '')}",
                '채널링크': f"https://www.youtube.com/channel/{snippet.get('channelId', '')}"
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _apply_main_sheet_formatting(self, workbook, worksheet, df):
        """메인 시트 포맷팅 적용"""
        # 포맷 정의
        formats = {
            'header': workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D7E4BC',
                'border': 1,
                'font_size': 11
            }),
            'number': workbook.add_format({'num_format': '#,##0'}),
            'percentage': workbook.add_format({'num_format': '0.00%'}),
            'url': workbook.add_format({'color': 'blue', 'underline': 1}),
            'shorts': workbook.add_format({'fg_color': '#FFE6E6'}),  # 쇼츠: 연한 빨강
            'long': workbook.add_format({'fg_color': '#E6F3FF'})     # 롱폼: 연한 파랑
        }
        
        # 헤더 포맷팅
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, formats['header'])
        
        # 행 높이 설정
        worksheet.set_default_row(config.THUMBNAIL_ROW_HEIGHT)
        
        # 데이터 행 포맷팅 (영상 유형별 색상)
        for row_num in range(1, len(df) + 1):
            video_type = df.iloc[row_num-1]['영상유형']
            
            if video_type == '쇼츠':
                worksheet.set_row(row_num, None, formats['shorts'])
            elif video_type == '롱폼':
                worksheet.set_row(row_num, None, formats['long'])
        
        # 열 너비 조정
        column_widths = {
            'A': 8,   # 순위
            'B': 15,  # 썸네일
            'C': 50,  # 제목
            'D': 20,  # 채널명
            'E': 15,  # 카테고리
            'F': 12,  # 영상유형
            'G': 12,  # 영상길이
            'H': 10,  # 영상길이_초
            'I': 20,  # 업로드일시
            'J': 12,  # 조회수
            'K': 10,  # 좋아요
            'L': 10,  # 댓글수
            'M': 12,  # 좋아요율
            'N': 12,  # 댓글율
            'O': 12,  # Outlier점수
            'P': 15,  # Outlier등급
            'Q': 12,  # 참여도점수
            'R': 15,  # 일평균조회수
            'S': 12,  # 성장속도
            'T': 15,  # 채널평균조회수
            'U': 15,  # 채널대비성과
            'V': 30,  # 핵심키워드
            'W': 10,  # 제목길이
            'X': 12,  # 댓글감정_긍정
            'Y': 12,  # 댓글감정_중립
            'Z': 12,  # 댓글감정_부정
            'AA': 40, # 영상링크
            'AB': 40  # 채널링크
        }
        
        for col, width in column_widths.items():
            worksheet.set_column(f'{col}:{col}', width)
        
        # 숫자 포맷 적용
        worksheet.set_column('J:L', 12, formats['number'])    # 조회수, 좋아요, 댓글수
        worksheet.set_column('M:N', 12, formats['percentage']) # 좋아요율, 댓글율
        worksheet.set_column('Q:Q', 12, formats['number'])    # 참여도점수
        worksheet.set_column('R:R', 15, formats['number'])    # 일평균조회수
        worksheet.set_column('T:T', 15, formats['number'])    # 채널평균조회수
        worksheet.set_column('AA:AB', 40, formats['url'])     # 링크들
        
        # 고정 창 설정
        worksheet.freeze_panes(1, 3)  # 헤더와 기본 정보 고정
    
    def _create_summary_sheet(self, writer, video_data_list, analysis_settings):
        """요약 정보 시트 생성"""
        workbook = writer.book
        summary_sheet = workbook.add_worksheet('📊 분석 요약')
        
        # 포맷 정의
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'fg_color': '#4F81BD',
            'color': 'white',
            'align': 'center'
        })
        
        section_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'fg_color': '#95B3D3',
            'color': 'white'
        })
        
        normal_format = workbook.add_format({'font_size': 12})
        number_format = workbook.add_format({'num_format': '#,##0'})
        
        row = 0
        
        # 메인 제목
        summary_sheet.merge_range('A1:D1', 'YouTube 트렌드 분석 리포트', title_format)
        row += 3
        
        # 분석 설정 정보
        summary_sheet.write(row, 0, '⚙️ 분석 설정', section_format)
        row += 2
        
        settings_info = [
            ['분석 모드', analysis_settings.get('mode_name', 'Unknown')],
            ['분석 지역', analysis_settings.get('region_name', 'Unknown')],
            ['영상 유형', analysis_settings.get('video_type_name', 'Unknown')],
            ['분석 영상 수', len(video_data_list)],
            ['분석 일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        # 키워드 검색 모드인 경우 추가 정보
        if analysis_settings.get('mode') == 'keyword':
            keyword_info = [
                ['검색 키워드', f"'{analysis_settings.get('keyword', 'Unknown')}'"],
                ['검색 기간', f"{analysis_settings.get('period_days', 'Unknown')}일"],
                ['최대 구독자', analysis_settings.get('max_subscribers_name', 'Unknown')],
                ['최소 조회수', analysis_settings.get('min_views_name', 'Unknown')]
            ]
            settings_info.extend(keyword_info)
        
        for info in settings_info:
            summary_sheet.write(row, 0, info[0], normal_format)
            summary_sheet.write(row, 1, info[1], normal_format)
            row += 1
        
        row += 2
        
        # 전체 통계 요약
        summary_sheet.write(row, 0, '📈 전체 통계', section_format)
        row += 2
        
        # 통계 계산
        total_views = sum(int(v.get('statistics', {}).get('viewCount', 0)) for v in video_data_list)
        total_likes = sum(int(v.get('statistics', {}).get('likeCount', 0)) for v in video_data_list)
        total_comments = sum(int(v.get('statistics', {}).get('commentCount', 0)) for v in video_data_list)
        avg_engagement = sum(v.get('analysis', {}).get('engagement_score', 0) for v in video_data_list) / len(video_data_list) if video_data_list else 0
        
        stats_info = [
            ['총 조회수', total_views],
            ['총 좋아요', total_likes],
            ['총 댓글수', total_comments],
            ['평균 참여도 점수', round(avg_engagement, 2)]
        ]
        
        for info in stats_info:
            summary_sheet.write(row, 0, info[0], normal_format)
            if isinstance(info[1], (int, float)) and info[0] != '평균 참여도 점수':
                summary_sheet.write(row, 1, info[1], number_format)
            else:
                summary_sheet.write(row, 1, info[1], normal_format)
            row += 1
        
        row += 2
        
        # 영상 유형별 분석
        summary_sheet.write(row, 0, '🎬 영상 유형별 분석', section_format)
        row += 2
        
        # 영상 유형별 통계 계산
        video_type_stats = self._calculate_video_type_stats(video_data_list)
        
        # 헤더
        headers = ['영상 유형', '개수', '총 조회수', '평균 조회수', '평균 Outlier점수']
        for col, header in enumerate(headers):
            summary_sheet.write(row, col, header, section_format)
        row += 1
        
        # 데이터
        for video_type, stats in video_type_stats.items():
            data_row = [
                video_type,
                stats['count'],
                stats['total_views'],
                stats['avg_views'],
                round(stats['avg_outlier'], 2)
            ]
            
            for col, value in enumerate(data_row):
                if col in [2, 3]:  # 조회수 컬럼
                    summary_sheet.write(row, col, value, number_format)
                else:
                    summary_sheet.write(row, col, value, normal_format)
            row += 1
        
        # 열 너비 조정
        summary_sheet.set_column('A:A', 25)
        summary_sheet.set_column('B:E', 15)
    
    def _create_keywords_sheet(self, writer, video_data_list):
        """트렌드 키워드 시트 생성"""
        try:
            from data import TextAnalyzer
            
            analyzer = TextAnalyzer()
            all_titles = [v.get('snippet', {}).get('title', '') for v in video_data_list]
            trending_keywords = analyzer.extract_trending_keywords(all_titles, max_keywords=50)
            
            # 키워드 빈도 계산
            keyword_freq = analyzer.analyze_keyword_frequency(all_titles)
            
            # 데이터프레임 생성
            keyword_data = []
            for i, keyword in enumerate(trending_keywords, 1):
                # 해당 키워드의 상세 정보 찾기
                keyword_detail = None
                for kw_info in keyword_freq.get('top_keywords', []):
                    if kw_info['keyword'] == keyword:
                        keyword_detail = kw_info
                        break
                
                if keyword_detail:
                    keyword_data.append({
                        '순위': i,
                        '키워드': keyword,
                        '출현 빈도': keyword_detail['count'],
                        '전체 대비 비율': f"{keyword_detail['frequency']:.1f}%",
                        '제목 포함률': f"{keyword_detail['title_coverage']:.1f}%",
                        '대표 영상': keyword_detail['containing_titles'][0]['title'][:30] + '...' if keyword_detail['containing_titles'] else ''
                    })
                else:
                    keyword_data.append({
                        '순위': i,
                        '키워드': keyword,
                        '출현 빈도': 1,
                        '전체 대비 비율': '0.1%',
                        '제목 포함률': '0.1%',
                        '대표 영상': ''
                    })
            
            df_keywords = pd.DataFrame(keyword_data)
            df_keywords.to_excel(writer, sheet_name='🔥 트렌드 키워드', index=False)
            
            # 포맷팅
            workbook = writer.book
            worksheet = writer.sheets['🔥 트렌드 키워드']
            
            header_format = workbook.add_format({
                'bold': True,
                'fg_color': '#FF6B35',
                'color': 'white',
                'border': 1
            })
            
            for col_num, value in enumerate(df_keywords.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            worksheet.set_column('A:A', 8)   # 순위
            worksheet.set_column('B:B', 20)  # 키워드
            worksheet.set_column('C:C', 12)  # 출현 빈도
            worksheet.set_column('D:D', 15)  # 전체 대비 비율
            worksheet.set_column('E:E', 15)  # 제목 포함률
            worksheet.set_column('F:F', 40)  # 대표 영상
            
        except ImportError:
            print("⚠️ 텍스트 분석 모듈을 가져올 수 없어 키워드 시트를 생략합니다.")
        except Exception as e:
            print(f"⚠️ 키워드 시트 생성 오류: {e}")
    
    def _create_video_type_analysis_sheet(self, writer, video_data_list):
        """영상 유형별 상세 분석 시트"""
        # 영상 유형별로 분리
        shorts_videos = [v for v in video_data_list if v.get('analysis', {}).get('video_type') == '쇼츠']
        long_videos = [v for v in video_data_list if v.get('analysis', {}).get('video_type') == '롱폼']
        
        if shorts_videos:
            self._create_type_specific_sheet(writer, shorts_videos, '🎬 쇼츠 분석')
        
        if long_videos:
            self._create_type_specific_sheet(writer, long_videos, '📹 롱폼 분석')
    
    def _create_performance_analysis_sheet(self, writer, video_data_list):
        """성과 분석 시트 생성"""
        workbook = writer.book
        perf_sheet = workbook.add_worksheet('🏆 성과 분석')
        
        # 성과별로 영상 분류
        high_performers = []  # Outlier Score 3.0 이상
        good_performers = []  # Outlier Score 1.5-3.0
        avg_performers = []   # Outlier Score 0.7-1.5
        poor_performers = []  # Outlier Score 0.7 미만
        
        for video in video_data_list:
            outlier_score = video.get('analysis', {}).get('outlier_score', 1.0)
            
            if outlier_score >= 3.0:
                high_performers.append(video)
            elif outlier_score >= 1.5:
                good_performers.append(video)
            elif outlier_score >= 0.7:
                avg_performers.append(video)
            else:
                poor_performers.append(video)
        
        # 성과 분포 요약
        total_videos = len(video_data_list)
        performance_summary = [
            ['성과 등급', '영상 수', '비율', '대표 영상'],
            ['🔥 고성과 (3.0x+)', len(high_performers), f"{len(high_performers)/total_videos*100:.1f}%", 
             high_performers[0]['snippet']['title'][:30] + '...' if high_performers else 'N/A'],
            ['📈 양호 (1.5-3.0x)', len(good_performers), f"{len(good_performers)/total_videos*100:.1f}%",
             good_performers[0]['snippet']['title'][:30] + '...' if good_performers else 'N/A'],
            ['😐 평균 (0.7-1.5x)', len(avg_performers), f"{len(avg_performers)/total_videos*100:.1f}%",
             avg_performers[0]['snippet']['title'][:30] + '...' if avg_performers else 'N/A'],
            ['📉 저조 (0.7x-)', len(poor_performers), f"{len(poor_performers)/total_videos*100:.1f}%",
             poor_performers[0]['snippet']['title'][:30] + '...' if poor_performers else 'N/A']
        ]
        
        # 데이터 쓰기
        for row, data in enumerate(performance_summary):
            for col, value in enumerate(data):
                perf_sheet.write(row, col, value)
        
        # 열 너비 조정
        perf_sheet.set_column('A:A', 20)
        perf_sheet.set_column('B:C', 12)
        perf_sheet.set_column('D:D', 40)
    
    def _create_type_specific_sheet(self, writer, videos, sheet_name):
        """특정 영상 유형 상세 분석 시트"""
        workbook = writer.book
        sheet = workbook.add_worksheet(sheet_name)
        
        # 포맷 정의
        header_format = workbook.add_format({
            'bold': True,
            'fg_color': '#4F81BD',
            'color': 'white'
        })
        
        # 기본 통계
        row = 0
        sheet.write(row, 0, f'{sheet_name} 상세 분석', header_format)
        row += 2
        
        # 상위 10개 영상
        sheet.write(row, 0, f'🏆 상위 10개 영상 (Outlier Score 기준)', header_format)
        row += 1
        
        headers = ['순위', '제목', '채널', '조회수', 'Outlier점수', '길이', '참여도']
        for col, header in enumerate(headers):
            sheet.write(row, col, header, header_format)
        row += 1
        
        # 상위 10개 데이터
        top_videos = sorted(videos, key=lambda x: x.get('analysis', {}).get('outlier_score', 0), reverse=True)[:10]
        
        for i, video in enumerate(top_videos, 1):
            snippet = video.get('snippet', {})
            statistics = video.get('statistics', {})
            analysis = video.get('analysis', {})
            
            data = [
                i,
                snippet.get('title', '')[:40] + '...' if len(snippet.get('title', '')) > 40 else snippet.get('title', ''),
                snippet.get('channelTitle', ''),
                int(statistics.get('viewCount', 0)),
                analysis.get('outlier_score', 1.0),
                analysis.get('formatted_duration', '00:00'),
                analysis.get('engagement_score', 0)
            ]
            
            for col, value in enumerate(data):
                sheet.write(row, col, value)
            row += 1
        
        # 열 너비 조정
        sheet.set_column('A:A', 8)
        sheet.set_column('B:B', 45)
        sheet.set_column('C:C', 20)
        sheet.set_column('D:D', 12)
        sheet.set_column('E:E', 12)
        sheet.set_column('F:F', 10)
        sheet.set_column('G:G', 12)
    
    def _insert_thumbnails(self, video_data_list):
        """썸네일 이미지를 엑셀에 삽입"""
        try:
            workbook = load_workbook(self.filename)
            worksheet = workbook['영상 분석 결과']
            
            thumbnails_found = False
            
            for i, video_data in enumerate(video_data_list, start=2):
                video_id = video_data.get('id', '')
                
                # 여러 가능한 썸네일 파일명 패턴 확인
                import glob
                possible_patterns = [
                    f'thumbnails/{i-1:03d}_*_{video_id}.jpg',
                    f'thumbnails/*_{video_id}.jpg',
                    f'thumbnails/{video_id}.jpg'
                ]
                
                thumbnail_path = None
                for pattern in possible_patterns:
                    matches = glob.glob(pattern)
                    if matches:
                        thumbnail_path = matches[0]
                        break
                
                if thumbnail_path and os.path.exists(thumbnail_path):
                    try:
                        img = image.Image(thumbnail_path)
                        img.width = 120
                        img.height = 68
                        
                        cell = f'B{i}'
                        worksheet.add_image(img, cell)
                        thumbnails_found = True
                        
                    except Exception as e:
                        worksheet[f'B{i}'] = '썸네일 오류'
                else:
                    worksheet[f'B{i}'] = '썸네일 없음'
            
            workbook.save(self.filename)
            
            if thumbnails_found:
                print("✅ 썸네일 이미지가 엑셀에 삽입되었습니다.")
            else:
                print("ℹ️ 삽입할 썸네일 이미지가 없습니다.")
            
        except Exception as e:
            print(f"⚠️ 썸네일 삽입 오류: {e}")
    
    def _add_charts(self, video_data_list):
        """차트 추가"""
        try:
            workbook = load_workbook(self.filename)
            
            # 차트 시트 생성
            if '📊 차트 분석' in workbook.sheetnames:
                chart_sheet = workbook['📊 차트 분석']
            else:
                chart_sheet = workbook.create_sheet('📊 차트 분석')
            
            # 영상 유형별 분포 차트
            self._create_video_type_pie_chart(workbook, chart_sheet, video_data_list)
            
            # 성과 분포 막대 차트
            self._create_performance_bar_chart(workbook, chart_sheet, video_data_list)
            
            workbook.save(self.filename)
            print("✅ 차트가 추가되었습니다.")
            
        except Exception as e:
            print(f"⚠️ 차트 추가 오류: {e}")
    
    def _create_video_type_pie_chart(self, workbook, sheet, video_data_list):
        """영상 유형별 파이 차트 생성"""
        try:
            # 데이터 준비
            type_counts = {'쇼츠': 0, '롱폼': 0, '기타': 0}
            for video in video_data_list:
                video_type = video.get('analysis', {}).get('video_type', '기타')
                type_counts[video_type] = type_counts.get(video_type, 0) + 1
            
            # 데이터 쓰기
            sheet['A1'] = '영상 유형'
            sheet['B1'] = '개수'
            
            row = 2
            for video_type, count in type_counts.items():
                if count > 0:
                    sheet[f'A{row}'] = video_type
                    sheet[f'B{row}'] = count
                    row += 1
            
            # 파이 차트 생성
            pie_chart = PieChart()
            labels = Reference(sheet, min_col=1, min_row=2, max_row=row-1)
            data = Reference(sheet, min_col=2, min_row=2, max_row=row-1)
            
            pie_chart.add_data(data)
            pie_chart.set_categories(labels)
            pie_chart.title = "영상 유형별 분포"
            
            sheet.add_chart(pie_chart, "D2")
            
        except Exception as e:
            print(f"파이 차트 생성 오류: {e}")
    
    def _create_performance_bar_chart(self, workbook, sheet, video_data_list):
        """성과 분포 막대 차트 생성"""
        try:
            # 성과 등급별 카운트
            performance_counts = {'고성과': 0, '양호': 0, '평균': 0, '저조': 0}
            
            for video in video_data_list:
                outlier_score = video.get('analysis', {}).get('outlier_score', 1.0)
                
                if outlier_score >= 3.0:
                    performance_counts['고성과'] += 1
                elif outlier_score >= 1.5:
                    performance_counts['양호'] += 1
                elif outlier_score >= 0.7:
                    performance_counts['평균'] += 1
                else:
                    performance_counts['저조'] += 1
            
            # 데이터 쓰기
            sheet['A10'] = '성과 등급'
            sheet['B10'] = '영상 수'
            
            row = 11
            for grade, count in performance_counts.items():
                sheet[f'A{row}'] = grade
                sheet[f'B{row}'] = count
                row += 1
            
            # 막대 차트 생성
            bar_chart = BarChart()
            data = Reference(sheet, min_col=2, min_row=10, max_row=14)
            categories = Reference(sheet, min_col=1, min_row=11, max_row=14)
            
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(categories)
            bar_chart.title = "성과 등급별 영상 분포"
            bar_chart.x_axis.title = "성과 등급"
            bar_chart.y_axis.title = "영상 수"
            
            sheet.add_chart(bar_chart, "D10")
            
        except Exception as e:
            print(f"막대 차트 생성 오류: {e}")
    
    def _calculate_video_type_stats(self, video_data_list):
        """영상 유형별 통계 계산"""
        video_type_stats = {}
        
        for video in video_data_list:
            video_type = video.get('analysis', {}).get('video_type', '알수없음')
            
            if video_type not in video_type_stats:
                video_type_stats[video_type] = {
                    'count': 0,
                    'total_views': 0,
                    'total_outlier': 0
                }
            
            stats = video.get('statistics', {})
            analysis = video.get('analysis', {})
            
            video_type_stats[video_type]['count'] += 1
            video_type_stats[video_type]['total_views'] += int(stats.get('viewCount', 0))
            video_type_stats[video_type]['total_outlier'] += analysis.get('outlier_score', 1.0)
        
        # 평균 계산
        for video_type in video_type_stats:
            count = video_type_stats[video_type]['count']
            if count > 0:
                video_type_stats[video_type]['avg_views'] = video_type_stats[video_type]['total_views'] // count
                video_type_stats[video_type]['avg_outlier'] = video_type_stats[video_type]['total_outlier'] / count
        
        return video_type_stats
    
    def _define_styles(self):
        """스타일 정의"""
        return {
            'header_style': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            }
        }
    
    def _format_datetime(self, datetime_str):
        """datetime 문자열을 읽기 쉬운 형태로 변환"""
        try:
            dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d %H:%M')
        except:
            return datetime_str
    
    def get_filename(self):
        """생성된 파일명 반환"""
        return self.filename


# 편의 함수들
def quick_excel_export(video_data_list, analysis_settings, filename=None):
    """
    빠른 엑셀 내보내기
    
    Args:
        video_data_list (list): 영상 데이터 목록
        analysis_settings (dict): 분석 설정
        filename (str): 파일명 (선택사항)
        
    Returns:
        str: 생성된 파일명
    """
    exporter = ExcelExporter(filename)
    return exporter.export_video_analysis(video_data_list, analysis_settings)

def export_comparison_report(video_groups, group_names, filename=None):
    """
    여러 그룹 비교 리포트 생성
    
    Args:
        video_groups (list): 영상 그룹들의 리스트
        group_names (list): 그룹명 리스트
        filename (str): 파일명 (선택사항)
        
    Returns:
        str: 생성된 파일명
    """
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"YouTube_Comparison_{timestamp}.xlsx"
    
    with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
        workbook = writer.book
        
        for i, (group, name) in enumerate(zip(video_groups, group_names)):
            # 각 그룹을 별도 시트로
            exporter = ExcelExporter()
            df = exporter._create_video_dataframe(group)
            
            sheet_name = f"{name} ({len(group)}개)"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            worksheet = writer.sheets[sheet_name]
            exporter._apply_main_sheet_formatting(workbook, worksheet, df)
    
    print(f"✅ 비교 리포트 생성 완료: {filename}")
    return filename