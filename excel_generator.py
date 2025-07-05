"""
엑셀 파일 생성 및 포매팅 관련 함수들 (수정된 버전)
"""

import os
import pandas as pd
from datetime import datetime
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.drawing import image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import config

class ExcelGenerator:
    def __init__(self, filename=None):
        """
        엑셀 생성기 초기화
        
        Args:
            filename (str): 생성할 엑셀 파일명
        """
        if filename:
            self.filename = filename
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.filename = f"YouTube_Trend_Analysis_{timestamp}.xlsx"
        
        self.workbook = None
        self.worksheet = None
    
    def create_excel_file(self, video_data_list, analysis_settings):
        """
        엑셀 파일 생성
        
        Args:
            video_data_list (list): 분석된 영상 데이터 목록
            analysis_settings (dict): 분석 설정 정보
        """
        try:
            # 데이터프레임 생성
            df = self._create_dataframe(video_data_list)
            
            # 엑셀 파일 생성 (xlsxwriter 사용)
            with pd.ExcelWriter(self.filename, engine='xlsxwriter') as writer:
                # 메인 데이터 시트
                df.to_excel(writer, sheet_name='영상 분석 결과', index=False)
                
                # 워크북과 워크시트 객체 가져오기
                workbook = writer.book
                worksheet = writer.sheets['영상 분석 결과']
                
                # 포매팅 적용
                self._apply_formatting(workbook, worksheet, df)
                
                # 요약 시트 생성
                self._create_summary_sheet(writer, video_data_list, analysis_settings)
                
                # 트렌드 키워드 시트 생성
                self._create_keywords_sheet(writer, video_data_list)
                
                # 영상 유형별 분석 시트 생성
                self._create_video_type_analysis_sheet(writer, video_data_list, analysis_settings)
            
            # 썸네일 이미지 삽입 (openpyxl 사용)
            self._insert_thumbnails(video_data_list)
            
            print(f"엑셀 파일이 생성되었습니다: {self.filename}")
            
        except Exception as e:
            print(f"엑셀 파일 생성 오류: {e}")
    
    def _create_dataframe(self, video_data_list):
        """영상 데이터를 데이터프레임으로 변환 - 개선된 버전"""
        data = []
        
        for video_data in video_data_list:
            snippet = video_data.get('snippet', {})
            statistics = video_data.get('statistics', {})
            content_details = video_data.get('contentDetails', {})
            analysis = video_data.get('analysis', {})
            
            # 영상 길이 정보 추가
            duration_seconds = 0
            try:
                from youtube_api import YouTubeAPIClient
                api_client = YouTubeAPIClient()
                duration_seconds = api_client.parse_duration(content_details.get('duration', 'PT0S'))
            except:
                duration_seconds = analysis.get('duration_seconds', 0)
            
            # 영상 유형 상세 정보
            video_type = analysis.get('video_type', '알수없음')
            formatted_duration = analysis.get('formatted_duration', '00:00')
            
            row = {
                '순위': video_data.get('rank', 0),
                '썸네일': '',  # 이미지는 별도로 삽입
                '제목': snippet.get('title', ''),
                '채널명': snippet.get('channelTitle', ''),
                '영상유형': video_type,
                '영상길이': formatted_duration,
                '영상길이_초': duration_seconds,
                '조회수': int(statistics.get('viewCount', 0)),
                '좋아요': int(statistics.get('likeCount', 0)),
                '댓글수': int(statistics.get('commentCount', 0)),
                'Outlier점수': analysis.get('outlier_score', 1.0),
                'Outlier등급': analysis.get('outlier_category', '😐 평균'),
                '채널평균조회수': analysis.get('channel_avg_views', 0),
                '업로드일시': self._format_datetime(snippet.get('publishedAt', '')),
                '카테고리': config.YOUTUBE_CATEGORIES.get(snippet.get('categoryId', ''), '기타'),
                '핵심키워드': ', '.join(analysis.get('keywords', [])),
                '댓글감정_긍정': f"{analysis.get('sentiment', {}).get('positive', 0)}%",
                '댓글감정_중립': f"{analysis.get('sentiment', {}).get('neutral', 0)}%",
                '댓글감정_부정': f"{analysis.get('sentiment', {}).get('negative', 0)}%",
                '참여도점수': analysis.get('engagement_score', 0),
                '일평균조회수': analysis.get('views_per_day', 0),
                '영상링크': f"https://www.youtube.com/watch?v={video_data.get('id', '')}"
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _apply_formatting(self, workbook, worksheet, df):
        """엑셀 시트 포매팅 적용"""
        # 포맷 정의
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D7E4BC',
            'border': 1
        })
        
        number_format = workbook.add_format({'num_format': '#,##0'})
        percent_format = workbook.add_format({'num_format': '0.00%'})
        url_format = workbook.add_format({'color': 'blue', 'underline': 1})
        
        # 영상 유형별 색상 포맷
        shorts_format = workbook.add_format({'fg_color': '#FFE6E6'})  # 연한 빨강
        long_format = workbook.add_format({'fg_color': '#E6F3FF'})   # 연한 파랑
        
        # 헤더 포매팅
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        
        # 행 높이 설정 (썸네일 때문에) - 수정된 부분
        worksheet.set_default_row(config.THUMBNAIL_ROW_HEIGHT)  # 기본 행 높이
        
        # 데이터 행 포매팅 (영상 유형별 색상)
        for row_num in range(1, len(df) + 1):
            worksheet.set_row(row_num, 80)
            video_type = df.iloc[row_num-1]['영상유형']
            row_format = shorts_format if video_type == '쇼츠' else long_format if video_type == '롱폼' else None
            
            if row_format:
                worksheet.set_row(row_num, None, row_format)
        
        # 열 너비 조정
        column_widths = {
            'A': 8,   # 순위
            'B': 15,  # 썸네일
            'C': 50,  # 제목
            'D': 20,  # 채널명
            'E': 12,  # 영상유형
            'F': 12,  # 영상길이
            'G': 10,  # 영상길이_초
            'H': 12,  # 조회수
            'I': 10,  # 좋아요
            'J': 10,  # 댓글수
            'K': 12,  # Outlier점수
            'L': 15,  # Outlier등급
            'M': 15,  # 채널평균조회수
            'N': 20,  # 업로드일시
            'O': 15,  # 카테고리
            'P': 30,  # 핵심키워드
            'Q': 12,  # 댓글감정_긍정
            'R': 12,  # 댓글감정_중립
            'S': 12,  # 댓글감정_부정
            'T': 12,  # 참여도점수
            'U': 15,  # 일평균조회수
            'V': 40   # 영상링크
        }
        
        for col, width in column_widths.items():
            worksheet.set_column(f'{col}:{col}', width)
        
        # 숫자 포맷 적용
        worksheet.set_column('H:J', 12, number_format)  # 조회수, 좋아요, 댓글수
        worksheet.set_column('M:M', 15, number_format)  # 채널평균조회수
        worksheet.set_column('T:U', 12, number_format)  # 참여도점수, 일평균조회수
        worksheet.set_column('V:V', 40, url_format)     # 영상링크
        
        # 고정 창 설정
        worksheet.freeze_panes(1, 0)
    
    def _create_summary_sheet(self, writer, video_data_list, analysis_settings):
        """요약 정보 시트 생성 - 개선된 버전"""
        workbook = writer.book
        summary_sheet = workbook.add_worksheet('분석 요약')
        
        # 제목 포맷
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'fg_color': '#4F81BD',
            'color': 'white'
        })
        
        # 일반 포맷
        normal_format = workbook.add_format({'font_size': 12})
        number_format = workbook.add_format({'num_format': '#,##0'})
        
        row = 0
        
        # 분석 설정 정보
        summary_sheet.write(row, 0, '📊 분석 설정 정보', title_format)
        row += 2
        
        settings_info = [
            ['분석 모드', analysis_settings.get('mode_name', 'Unknown')],
            ['분석 지역', analysis_settings.get('region_name', 'Unknown')],
            ['영상 유형', analysis_settings.get('video_type_name', 'Unknown')],
            ['분석 영상 수', len(video_data_list)],
            ['분석 일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        # 키워드 검색 모드인 경우 추가 정보
        if analysis_settings.get('mode') == 'keyword':
            keyword_info = [
                ['검색 키워드', f"'{analysis_settings.get('keyword', 'Unknown')}'"],
                ['검색 기간', f"{analysis_settings.get('period_days', 'Unknown')}일"],
                ['최대 구독자', analysis_settings.get('max_subscribers_name', 'Unknown')],
                ['최소 조회수', analysis_settings.get('min_views_name', 'Unknown')]
            ]
            settings_info.extend(keyword_info)
        else:
            # 트렌딩 모드인 경우
            settings_info.insert(2, ['카테고리', analysis_settings.get('category_name', 'Unknown')])
        
        for info in settings_info:
            summary_sheet.write(row, 0, info[0], normal_format)
            summary_sheet.write(row, 1, info[1], normal_format)
            row += 1
        
        row += 2
        
        # 영상 유형별 분석
        summary_sheet.write(row, 0, '🎬 영상 유형별 분석', title_format)
        row += 2
        
        # 영상 유형별 통계 계산
        video_type_stats = {}
        for video in video_data_list:
            video_type = video.get('analysis', {}).get('video_type', '알수없음')
            if video_type not in video_type_stats:
                video_type_stats[video_type] = {
                    'count': 0,
                    'total_views': 0,
                    'total_likes': 0,
                    'avg_outlier': 0
                }
            
            stats = video.get('statistics', {})
            analysis = video.get('analysis', {})
            
            video_type_stats[video_type]['count'] += 1
            video_type_stats[video_type]['total_views'] += int(stats.get('viewCount', 0))
            video_type_stats[video_type]['total_likes'] += int(stats.get('likeCount', 0))
            video_type_stats[video_type]['avg_outlier'] += analysis.get('outlier_score', 1.0)
        
        # 평균 계산
        for video_type in video_type_stats:
            count = video_type_stats[video_type]['count']
            if count > 0:
                video_type_stats[video_type]['avg_outlier'] /= count
        
        # 영상 유형별 정보 출력
        type_info = [
            ['영상 유형', '개수', '총 조회수', '총 좋아요', '평균 Outlier점수']
        ]
        
        for video_type, stats in video_type_stats.items():
            type_info.append([
                video_type,
                stats['count'],
                stats['total_views'],
                stats['total_likes'],
                round(stats['avg_outlier'], 2)
            ])
        
        for i, info in enumerate(type_info):
            for j, value in enumerate(info):
                if i == 0:  # 헤더
                    summary_sheet.write(row + i, j, value, title_format)
                else:
                    if j in [2, 3]:  # 숫자 컬럼
                        summary_sheet.write(row + i, j, value, number_format)
                    else:
                        summary_sheet.write(row + i, j, value, normal_format)
        
        row += len(type_info) + 2
        
        # 전체 통계 요약
        summary_sheet.write(row, 0, '📈 전체 통계 요약', title_format)
        row += 2
        
        # 통계 계산
        total_views = sum(int(v.get('statistics', {}).get('viewCount', 0)) for v in video_data_list)
        total_likes = sum(int(v.get('statistics', {}).get('likeCount', 0)) for v in video_data_list)
        total_comments = sum(int(v.get('statistics', {}).get('commentCount', 0)) for v in video_data_list)
        avg_engagement = sum(v.get('analysis', {}).get('engagement_score', 0) for v in video_data_list) / len(video_data_list) if video_data_list else 0
        
        stats_info = [
            ['총 조회수', total_views],
            ['총 좋아요', total_likes],
            ['총 댓글수', total_comments],
            ['평균 참여도 점수', round(avg_engagement, 2)]
        ]
        
        for info in stats_info:
            summary_sheet.write(row, 0, info[0], normal_format)
            if isinstance(info[1], (int, float)) and info[0] != '평균 참여도 점수':
                summary_sheet.write(row, 1, info[1], number_format)
            else:
                summary_sheet.write(row, 1, info[1], normal_format)
            row += 1
        
        # 열 너비 조정
        summary_sheet.set_column('A:A', 20)
        summary_sheet.set_column('B:B', 20)
        summary_sheet.set_column('C:C', 15)
        summary_sheet.set_column('D:D', 15)
        summary_sheet.set_column('E:E', 15)
    
    def _create_video_type_analysis_sheet(self, writer, video_data_list, analysis_settings):
        """영상 유형별 상세 분석 시트 생성"""
        workbook = writer.book
        
        # 영상 유형별로 분리
        shorts_videos = [v for v in video_data_list if v.get('analysis', {}).get('video_type') == '쇼츠']
        long_videos = [v for v in video_data_list if v.get('analysis', {}).get('video_type') == '롱폼']
        
        if shorts_videos:
            self._create_type_specific_sheet(writer, shorts_videos, '쇼츠 분석', workbook)
        
        if long_videos:
            self._create_type_specific_sheet(writer, long_videos, '롱폼 분석', workbook)
    
    def _create_type_specific_sheet(self, writer, videos, sheet_name, workbook):
        """특정 영상 유형에 대한 상세 분석 시트"""
        sheet = workbook.add_worksheet(sheet_name)
        
        # 포맷 정의
        header_format = workbook.add_format({
            'bold': True,
            'fg_color': '#4F81BD',
            'color': 'white'
        })
        
        # 기본 통계
        row = 0
        sheet.write(row, 0, f'📊 {sheet_name} 상세 분석', header_format)
        row += 2
        
        # 상위 10개 영상
        sheet.write(row, 0, f'🏆 상위 10개 {sheet_name.split()[0]} 영상', header_format)
        row += 1
        
        headers = ['순위', '제목', '채널', '조회수', 'Outlier점수', '길이']
        for col, header in enumerate(headers):
            sheet.write(row, col, header, header_format)
        row += 1
        
        # 상위 10개 데이터
        top_videos = sorted(videos, key=lambda x: x.get('analysis', {}).get('outlier_score', 0), reverse=True)[:10]
        
        for i, video in enumerate(top_videos, 1):
            snippet = video.get('snippet', {})
            statistics = video.get('statistics', {})
            analysis = video.get('analysis', {})
            
            data = [
                i,
                snippet.get('title', '')[:40] + '...' if len(snippet.get('title', '')) > 40 else snippet.get('title', ''),
                snippet.get('channelTitle', ''),
                int(statistics.get('viewCount', 0)),
                analysis.get('outlier_score', 1.0),
                analysis.get('formatted_duration', '00:00')
            ]
            
            for col, value in enumerate(data):
                sheet.write(row, col, value)
            row += 1
        
        # 열 너비 조정
        sheet.set_column('A:A', 8)
        sheet.set_column('B:B', 45)
        sheet.set_column('C:C', 20)
        sheet.set_column('D:D', 12)
        sheet.set_column('E:E', 12)
        sheet.set_column('F:F', 10)
    
    def _create_keywords_sheet(self, writer, video_data_list):
        """트렌드 키워드 시트 생성"""
        from data_analyzer import DataAnalyzer
        
        analyzer = DataAnalyzer()
        all_titles = [v.get('snippet', {}).get('title', '') for v in video_data_list]
        trending_keywords = analyzer.extract_trending_keywords(all_titles, max_keywords=50)
        
        # 키워드 빈도 계산
        from collections import Counter
        all_keywords = []
        for title in all_titles:
            keywords = analyzer.extract_keywords_from_title(title)
            all_keywords.extend(keywords)
        
        keyword_counts = Counter(all_keywords)
        
        # 데이터프레임 생성
        keyword_data = []
        for i, keyword in enumerate(trending_keywords, 1):
            keyword_data.append({
                '순위': i,
                '키워드': keyword,
                '출현 빈도': keyword_counts.get(keyword, 0),
                '비율': f"{(keyword_counts.get(keyword, 0) / len(video_data_list) * 100):.1f}%"
            })
        
        df_keywords = pd.DataFrame(keyword_data)
        df_keywords.to_excel(writer, sheet_name='트렌드 키워드', index=False)
        
        # 포매팅
        workbook = writer.book
        worksheet = writer.sheets['트렌드 키워드']
        
        header_format = workbook.add_format({
            'bold': True,
            'fg_color': '#D7E4BC',
            'border': 1
        })
        
        for col_num, value in enumerate(df_keywords.columns.values):
            worksheet.write(0, col_num, value, header_format)
        
        worksheet.set_column('A:A', 8)   # 순위
        worksheet.set_column('B:B', 20)  # 키워드
        worksheet.set_column('C:C', 12)  # 출현 빈도
        worksheet.set_column('D:D', 10)  # 비율
    
    def _insert_thumbnails(self, video_data_list):
        """썸네일 이미지를 엑셀에 삽입 (선택적 다운로드 지원)"""
        try:
            workbook = load_workbook(self.filename)
            worksheet = workbook['영상 분석 결과']
            
            # 썸네일 폴더가 있는지 확인
            thumbnails_found = False
            
            for i, video_data in enumerate(video_data_list, start=2):  # 2행부터 시작 (1행은 헤더)
                video_id = video_data.get('id', '')
                
                # 여러 가능한 썸네일 파일명 패턴 확인
                possible_patterns = [
                    f'thumbnails/{i-1:03d}_*_{video_id}.jpg',  # 순위_제목_ID.jpg
                    f'thumbnails/*_{video_id}.jpg',            # 제목_ID.jpg
                    f'thumbnails/{video_id}.jpg'               # ID.jpg
                ]
                
                thumbnail_path = None
                import glob
                
                for pattern in possible_patterns:
                    matches = glob.glob(pattern)
                    if matches:
                        thumbnail_path = matches[0]
                        break
                
                if thumbnail_path and os.path.exists(thumbnail_path):
                    try:
                        # 이미지 로드 및 리사이즈
                        img = image.Image(thumbnail_path)
                        
                        # 썸네일 크기 조정 (셀 크기에 맞게)
                        img.width = 120
                        img.height = 68
                        
                        # B열(썸네일 열)에 이미지 삽입
                        cell = f'B{i}'
                        worksheet.add_image(img, cell)
                        
                        thumbnails_found = True
                        
                    except Exception as e:
                        print(f"썸네일 삽입 오류 (영상 ID: {video_id}): {e}")
                        worksheet[f'B{i}'] = '썸네일 오류'
                else:
                    # 썸네일 파일이 없는 경우
                    worksheet[f'B{i}'] = '썸네일 없음'
            
            workbook.save(self.filename)
            
            if thumbnails_found:
                print("✅ 일부 썸네일 이미지가 엑셀에 삽입되었습니다.")
            else:
                print("ℹ️ 삽입할 썸네일 이미지가 없습니다. 선택적 다운로드를 사용하세요.")
            
        except Exception as e:
            print(f"썸네일 삽입 전체 오류: {e}")
    
    def _format_datetime(self, datetime_str):
        """datetime 문자열을 읽기 쉬운 형태로 변환"""
        try:
            dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d %H:%M')
        except:
            return datetime_str
    
    def add_charts(self, video_data_list):
        """차트 추가 (선택적 기능)"""
        try:
            workbook = load_workbook(self.filename)
            
            # 차트 시트 생성
            chart_sheet = workbook.create_sheet('차트 분석')
            
            # 여기에 차트 생성 코드 추가 가능
            # (matplotlib 또는 openpyxl의 차트 기능 사용)
            
            workbook.save(self.filename)
            
        except Exception as e:
            print(f"차트 추가 오류: {e}")
    
    def get_filename(self):
        """생성된 파일명 반환"""
        return self.filename